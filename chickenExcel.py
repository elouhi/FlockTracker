#!/usr/bin/env/ python3
import openpyxl
import os

#Save WB New Name
def saveNewName(wBName, wb):
    wb.save(os.path.dirname(os.path.abspath(__file__))+"/"+wBName+".xlsx")
    print(os.path.dirname(os.path.abspath(__file__))+"/"+wBName+".xlsx")
	
def openWB(wBName):
    wb = load_workbook(os.path.dirname(os.path.abspath(__file__))+"/"+wBName+".xlsx")

  #Print WorkSheet
def printWSTerminal(wb):  
    sheet=wb.active
    for sheet in wb:          
        max_row=sheet.max_row
        max_column=sheet.max_column
        print(sheet.title+"\n")
        for i in range(1,sheet.max_row+1):
            for j in range(1,sheet.max_column+1):
                cellObj=sheet.cell(row=i,column=j)
                print(cellObj.value)
            print("\n")
    
def addSheet(wb, wBName):
    sheetName = input("Name of new Sheet: \n")
    try:
        wb.create_sheet(sheetName)
        wb.save(os.path.dirname(os.path.abspath(__file__))+"/"+wBName+".xlsx")
    except:
        print("Error could not access/save sheet")

def removeSheet(wb, wBName):
    sheetName = input("Name of Sheet to Remove: \n")
    try:
        wb.remove_sheet(wb.get_sheet_by_name(sheetName))
        wb.save(os.path.dirname(os.path.abspath(__file__))+"/"+wBName+".xlsx")
    except:
        print("Error could not access/save sheet")

	